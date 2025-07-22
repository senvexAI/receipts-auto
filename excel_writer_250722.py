# === 모듈 2: excel_writer.py ===
import os
from openpyxl import load_workbook

def read_text_files(output_text_folder):
    details_path = os.path.join(output_text_folder, "교통비내역.txt")
    summary_path = os.path.join(output_text_folder, "직원별합계.txt")

    details = []
    with open(details_path, "r", encoding="utf-8") as f:
        next(f)
        for line in f:
            details.append(line.strip().split("\t"))

    summary = {}
    with open(summary_path, "r", encoding="utf-8") as f:
        next(f)
        for line in f:
            name, amount = line.strip().split("\t")
            summary[name] = amount

    return details, summary

def write_to_excel(template_path, output_excel, details, summary):
    wb = load_workbook(template_path)

    # 교통비내역
    if "교통비내역" in wb.sheetnames:
        ws_details = wb["교통비내역"]
        row_idx = 2
        for row_data in details:
            for col_idx, value in enumerate(row_data, start=1):
                ws_details.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    # 직원별 사용금액
    if "직원별 사용금액" not in wb.sheetnames:
        ws_summary = wb.create_sheet("직원별 사용금액")
    else:
        ws_summary = wb["직원별 사용금액"]

    ws_summary.append(["적요", "직원명", "총합계"])
    for name, total in summary.items():
        ws_summary.append(["교통비", name, total])

    wb.save(output_excel)

def generate_excel(output_text_folder, template_path, output_excel, progress_callback=None):
    if progress_callback: progress_callback(75)  # Excel 시작
    details, summary = read_text_files(output_text_folder)
    write_to_excel(template_path, output_excel, details, summary)
    if progress_callback: progress_callback(100)  # Excel 완료
    return True